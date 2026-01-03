# -*- coding: utf-8 -*-
"""3業界パターンのテストデータ生成スクリプト

パターン:
1. 介護業界 - 社会福祉法人さくら会
2. 医療業界 - 医療法人青空会
3. 障害福祉業界 - NPO法人ひまわり

各パターンで担当者構成・職種・応募経路を明確に分離
"""

import csv
from datetime import datetime, timedelta
from pathlib import Path
import random

# 出力ディレクトリ
OUTPUT_DIR = Path(__file__).parent

# ========================================
# パターン1: 介護業界
# ========================================
KAIGO_CONFIG = {
    'tenant': {
        'company_name': '社会福祉法人さくら会',
        'tenant_code': 'sakura-kai',
        'notification_email': 'recruit@sakura-kai.or.jp',
    },
    'departments': ['介護部', '訪問介護事業部', '居宅支援事業部', '経営企画部'],
    'staff': [
        {'id': 'u001', 'name': '山本 一郎', 'role': '採用担当', 'dept': '経営企画部'},
        {'id': 'u002', 'name': '鈴木 美香', 'role': '介護部長', 'dept': '介護部'},
        {'id': 'u003', 'name': '佐藤 健一', 'role': '施設長（さくら苑）', 'dept': '介護部'},
        {'id': 'u004', 'name': '田中 由美', 'role': 'サービス提供責任者', 'dept': '訪問介護事業部'},
        {'id': 'u005', 'name': '高橋 誠', 'role': '居宅管理者', 'dept': '居宅支援事業部'},
    ],
    'jobs': [
        {'code': 'K-001', 'title': '介護福祉士（特別養護老人ホーム）', 'dept': '介護部', 'type': 'full_time', 'salary_min': 300, 'salary_max': 400, 'location': '東京都世田谷区', 'headcount': 3},
        {'code': 'K-002', 'title': 'ケアマネジャー（居宅介護支援）', 'dept': '居宅支援事業部', 'type': 'full_time', 'salary_min': 350, 'salary_max': 450, 'location': '東京都新宿区', 'headcount': 1},
        {'code': 'K-003', 'title': '訪問介護スタッフ（サ責候補）', 'dept': '訪問介護事業部', 'type': 'full_time', 'salary_min': 320, 'salary_max': 400, 'location': '東京都杉並区', 'headcount': 2},
        {'code': 'K-004', 'title': '介護職員（デイサービス）', 'dept': '介護部', 'type': 'full_time', 'salary_min': 250, 'salary_max': 320, 'location': '東京都練馬区', 'headcount': 5},
        {'code': 'K-005', 'title': '夜勤専従介護スタッフ', 'dept': '介護部', 'type': 'part_time', 'salary_min': None, 'salary_max': None, 'location': '東京都大田区', 'headcount': 3},
        {'code': 'K-006', 'title': '生活相談員', 'dept': '介護部', 'type': 'full_time', 'salary_min': 300, 'salary_max': 380, 'location': '東京都世田谷区', 'headcount': 1},
        {'code': 'K-007', 'title': '機能訓練指導員（PT/OT）', 'dept': '介護部', 'type': 'full_time', 'salary_min': 380, 'salary_max': 480, 'location': '神奈川県横浜市', 'headcount': 1},
        {'code': 'K-008', 'title': '管理栄養士', 'dept': '介護部', 'type': 'full_time', 'salary_min': 320, 'salary_max': 400, 'location': '東京都板橋区', 'headcount': 1},
    ],
    'candidates': [
        {'name': '山田 太郎', 'kana': 'やまだ たろう', 'gender': 'male', 'age': 35, 'exp': 10, 'current': '特養さくら園', 'position': 'ユニットリーダー', 'qual': '介護福祉士,認知症ケア専門士', 'desired': '介護職,リーダー職'},
        {'name': '鈴木 花子', 'kana': 'すずき はなこ', 'gender': 'female', 'age': 42, 'exp': 15, 'current': '有料老人ホームひまわり', 'position': '介護主任', 'qual': '介護福祉士,介護支援専門員', 'desired': 'ケアマネジャー'},
        {'name': '佐藤 健一', 'kana': 'さとう けんいち', 'gender': 'male', 'age': 38, 'exp': 12, 'current': '居宅介護支援事業所', 'position': '主任ケアマネ', 'qual': '介護支援専門員,主任介護支援専門員', 'desired': 'ケアマネジャー,管理職'},
        {'name': '田中 美咲', 'kana': 'たなか みさき', 'gender': 'female', 'age': 28, 'exp': 5, 'current': '回復期リハビリ病院', 'position': '理学療法士', 'qual': '理学療法士', 'desired': 'リハビリ職'},
        {'name': '伊藤 翔太', 'kana': 'いとう しょうた', 'gender': 'male', 'age': 23, 'exp': 0, 'current': '無職（就職活動中）', 'position': '', 'qual': '介護職員初任者研修（取得予定）', 'desired': '介護職'},
        {'name': '高橋 由美', 'kana': 'たかはし ゆみ', 'gender': 'female', 'age': 45, 'exp': 18, 'current': '訪問介護ステーション', 'position': 'サービス提供責任者', 'qual': '介護福祉士,サービス提供責任者', 'desired': '訪問介護,サ責'},
        {'name': '渡辺 誠', 'kana': 'わたなべ まこと', 'gender': 'male', 'age': 32, 'exp': 8, 'current': 'グループホーム', 'position': '介護職員', 'qual': '介護福祉士', 'desired': '介護職,特養'},
        {'name': '中村 あやか', 'kana': 'なかむら あやか', 'gender': 'female', 'age': 29, 'exp': 6, 'current': '療養型病院', 'position': '准看護師', 'qual': '准看護師', 'desired': '看護職,施設'},
        {'name': '小林 拓也', 'kana': 'こばやし たくや', 'gender': 'male', 'age': 33, 'exp': 9, 'current': 'デイケア', 'position': '作業療法士', 'qual': '作業療法士,認知症リハ専門', 'desired': 'リハビリ職'},
        {'name': '加藤 恵美', 'kana': 'かとう えみ', 'gender': 'female', 'age': 40, 'exp': 14, 'current': '特別養護老人ホーム', 'position': '管理栄養士', 'qual': '管理栄養士,嚥下食指導', 'desired': '栄養士'},
        {'name': '松本 大輔', 'kana': 'まつもと だいすけ', 'gender': 'male', 'age': 27, 'exp': 4, 'current': '有料老人ホーム', 'position': '介護職員', 'qual': '介護福祉士', 'desired': '夜勤専従'},
        {'name': '吉田 さくら', 'kana': 'よしだ さくら', 'gender': 'female', 'age': 25, 'exp': 3, 'current': 'デイサービス', 'position': '介護職員', 'qual': '介護職員初任者研修', 'desired': 'デイサービス'},
        {'name': '斎藤 雄太', 'kana': 'さいとう ゆうた', 'gender': 'male', 'age': 36, 'exp': 10, 'current': '急性期病院', 'position': '看護師', 'qual': '正看護師', 'desired': '看護職,施設'},
        {'name': '森田 真理子', 'kana': 'もりた まりこ', 'gender': 'female', 'age': 44, 'exp': 16, 'current': '居宅介護支援事業所', 'position': 'ケアマネジャー', 'qual': '介護支援専門員,社会福祉士', 'desired': 'ケアマネジャー'},
        {'name': '木村 健太郎', 'kana': 'きむら けんたろう', 'gender': 'male', 'age': 30, 'exp': 6, 'current': '老健', 'position': '理学療法士', 'qual': '理学療法士', 'desired': 'リハビリ職'},
        {'name': '岡田 麻衣', 'kana': 'おかだ まい', 'gender': 'female', 'age': 22, 'exp': 0, 'current': '専門学校生', 'position': '', 'qual': '介護福祉士（取得予定）', 'desired': '介護職'},
        {'name': '石井 大樹', 'kana': 'いしい だいき', 'gender': 'male', 'age': 34, 'exp': 9, 'current': 'グループホーム', 'position': '介護リーダー', 'qual': '介護福祉士', 'desired': '介護職,特養'},
        {'name': '清水 優子', 'kana': 'しみず ゆうこ', 'gender': 'female', 'age': 48, 'exp': 20, 'current': '訪問看護ステーション', 'position': '管理者', 'qual': '正看護師,訪問看護認定', 'desired': '看護職,管理職'},
        {'name': '山本 拓海', 'kana': 'やまもと たくみ', 'gender': 'male', 'age': 31, 'exp': 7, 'current': '回復期病院', 'position': '言語聴覚士', 'qual': '言語聴覚士', 'desired': 'リハビリ職,ST'},
        {'name': '藤田 美穂', 'kana': 'ふじた みほ', 'gender': 'female', 'age': 24, 'exp': 1, 'current': '飲食店アルバイト', 'position': '', 'qual': '', 'desired': '介護職,未経験'},
    ],
    'sources': ['マイナビ介護', 'リクルート介護', 'カイゴジョブ', 'ハローワーク', '自社採用サイト', 'Indeed'],
}

# ========================================
# パターン2: 医療業界
# ========================================
IRYO_CONFIG = {
    'tenant': {
        'company_name': '医療法人青空会',
        'tenant_code': 'aozora-kai',
        'notification_email': 'jinji@aozora-hospital.or.jp',
    },
    'departments': ['看護部', '診療技術部', '事務部', '人事課'],
    'staff': [
        {'id': 'u001', 'name': '井上 直子', 'role': '人事課長', 'dept': '人事課'},
        {'id': 'u002', 'name': '中田 美智子', 'role': '看護部長', 'dept': '看護部'},
        {'id': 'u003', 'name': '大西 健二', 'role': '事務長', 'dept': '事務部'},
        {'id': 'u004', 'name': '藤井 恵子', 'role': '看護師長（外来）', 'dept': '看護部'},
        {'id': 'u005', 'name': '村上 正', 'role': '診療技術部長', 'dept': '診療技術部'},
    ],
    'jobs': [
        {'code': 'I-001', 'title': '正看護師（病棟）', 'dept': '看護部', 'type': 'full_time', 'salary_min': 400, 'salary_max': 550, 'location': '東京都中央区', 'headcount': 5},
        {'code': 'I-002', 'title': '正看護師（外来）', 'dept': '看護部', 'type': 'full_time', 'salary_min': 380, 'salary_max': 500, 'location': '東京都中央区', 'headcount': 3},
        {'code': 'I-003', 'title': '准看護師（透析室）', 'dept': '看護部', 'type': 'full_time', 'salary_min': 340, 'salary_max': 420, 'location': '東京都中央区', 'headcount': 2},
        {'code': 'I-004', 'title': '臨床検査技師', 'dept': '診療技術部', 'type': 'full_time', 'salary_min': 350, 'salary_max': 450, 'location': '東京都中央区', 'headcount': 2},
        {'code': 'I-005', 'title': '診療放射線技師', 'dept': '診療技術部', 'type': 'full_time', 'salary_min': 380, 'salary_max': 480, 'location': '東京都中央区', 'headcount': 1},
        {'code': 'I-006', 'title': '医療事務（受付）', 'dept': '事務部', 'type': 'full_time', 'salary_min': 250, 'salary_max': 320, 'location': '東京都中央区', 'headcount': 3},
        {'code': 'I-007', 'title': '医療事務（レセプト）', 'dept': '事務部', 'type': 'full_time', 'salary_min': 280, 'salary_max': 350, 'location': '東京都中央区', 'headcount': 2},
        {'code': 'I-008', 'title': '薬剤師', 'dept': '診療技術部', 'type': 'full_time', 'salary_min': 450, 'salary_max': 600, 'location': '東京都中央区', 'headcount': 1},
        {'code': 'I-009', 'title': '理学療法士（リハビリ科）', 'dept': '診療技術部', 'type': 'full_time', 'salary_min': 380, 'salary_max': 480, 'location': '東京都中央区', 'headcount': 2},
        {'code': 'I-010', 'title': '看護助手', 'dept': '看護部', 'type': 'part_time', 'salary_min': None, 'salary_max': None, 'location': '東京都中央区', 'headcount': 4},
    ],
    'candidates': [
        {'name': '西村 真由美', 'kana': 'にしむら まゆみ', 'gender': 'female', 'age': 32, 'exp': 9, 'current': '大学病院', 'position': '正看護師', 'qual': '正看護師,ICU経験', 'desired': '看護職,病棟'},
        {'name': '上田 洋介', 'kana': 'うえだ ようすけ', 'gender': 'male', 'age': 35, 'exp': 10, 'current': '総合病院', 'position': '臨床検査技師', 'qual': '臨床検査技師,超音波検査士', 'desired': '臨床検査'},
        {'name': '原田 美紀', 'kana': 'はらだ みき', 'gender': 'female', 'age': 28, 'exp': 5, 'current': 'クリニック', 'position': '正看護師', 'qual': '正看護師', 'desired': '看護職,外来'},
        {'name': '福田 健太', 'kana': 'ふくだ けんた', 'gender': 'male', 'age': 30, 'exp': 6, 'current': '急性期病院', 'position': '診療放射線技師', 'qual': '診療放射線技師,CT認定', 'desired': '放射線技師'},
        {'name': '長谷川 愛', 'kana': 'はせがわ あい', 'gender': 'female', 'age': 26, 'exp': 3, 'current': '調剤薬局', 'position': '薬剤師', 'qual': '薬剤師', 'desired': '薬剤師,病院'},
        {'name': '山口 誠一', 'kana': 'やまぐち せいいち', 'gender': 'male', 'age': 38, 'exp': 13, 'current': '透析クリニック', 'position': '正看護師', 'qual': '正看護師,透析技術認定士', 'desired': '看護職,透析'},
        {'name': '前田 恵子', 'kana': 'まえだ けいこ', 'gender': 'female', 'age': 45, 'exp': 20, 'current': '総合病院', 'position': '看護師長', 'qual': '正看護師,認定看護管理者', 'desired': '看護管理'},
        {'name': '小川 翼', 'kana': 'おがわ つばさ', 'gender': 'male', 'age': 27, 'exp': 4, 'current': 'リハビリ病院', 'position': '理学療法士', 'qual': '理学療法士', 'desired': 'リハビリ職'},
        {'name': '村田 優子', 'kana': 'むらた ゆうこ', 'gender': 'female', 'age': 33, 'exp': 8, 'current': '健診センター', 'position': '臨床検査技師', 'qual': '臨床検査技師,細胞検査士', 'desired': '臨床検査'},
        {'name': '近藤 大地', 'kana': 'こんどう だいち', 'gender': 'male', 'age': 29, 'exp': 5, 'current': '精神科病院', 'position': '正看護師', 'qual': '正看護師,精神科経験', 'desired': '看護職'},
        {'name': '遠藤 彩', 'kana': 'えんどう あや', 'gender': 'female', 'age': 25, 'exp': 2, 'current': 'クリニック', 'position': '医療事務', 'qual': '医療事務技能審査', 'desired': '医療事務'},
        {'name': '内田 俊介', 'kana': 'うちだ しゅんすけ', 'gender': 'male', 'age': 31, 'exp': 7, 'current': '整形外科', 'position': '診療放射線技師', 'qual': '診療放射線技師,MRI専門', 'desired': '放射線技師'},
        {'name': '後藤 麻衣', 'kana': 'ごとう まい', 'gender': 'female', 'age': 24, 'exp': 1, 'current': '歯科医院', 'position': '受付事務', 'qual': '医療事務資格', 'desired': '医療事務'},
        {'name': '杉山 正人', 'kana': 'すぎやま まさと', 'gender': 'male', 'age': 42, 'exp': 17, 'current': '大学病院', 'position': '薬剤師', 'qual': '薬剤師,がん専門薬剤師', 'desired': '薬剤師'},
        {'name': '宮崎 千尋', 'kana': 'みやざき ちひろ', 'gender': 'female', 'age': 30, 'exp': 6, 'current': '産婦人科', 'position': '助産師', 'qual': '助産師,正看護師', 'desired': '助産師,看護職'},
        {'name': '新井 拓真', 'kana': 'あらい たくま', 'gender': 'male', 'age': 26, 'exp': 3, 'current': 'リハビリ病院', 'position': '作業療法士', 'qual': '作業療法士', 'desired': 'リハビリ職'},
        {'name': '横山 美月', 'kana': 'よこやま みづき', 'gender': 'female', 'age': 23, 'exp': 0, 'current': '看護学校卒業', 'position': '', 'qual': '正看護師（新卒）', 'desired': '看護職,新卒'},
        {'name': '阿部 浩二', 'kana': 'あべ こうじ', 'gender': 'male', 'age': 48, 'exp': 22, 'current': '総合病院', 'position': '診療技術部長', 'qual': '臨床検査技師,管理職経験', 'desired': '管理職'},
        {'name': '岩田 理恵', 'kana': 'いわた りえ', 'gender': 'female', 'age': 36, 'exp': 12, 'current': '透析クリニック', 'position': '准看護師', 'qual': '准看護師', 'desired': '看護職,透析'},
        {'name': '谷口 翔', 'kana': 'たにぐち しょう', 'gender': 'male', 'age': 28, 'exp': 4, 'current': '訪問リハビリ', 'position': '言語聴覚士', 'qual': '言語聴覚士', 'desired': 'リハビリ職,ST'},
    ],
    'sources': ['マイナビ看護師', 'ナース人材バンク', 'レバウェル看護', 'ジョブメドレー', 'ハローワーク', '自社採用サイト'],
}

# ========================================
# パターン3: 障害福祉業界
# ========================================
SHOGAI_CONFIG = {
    'tenant': {
        'company_name': 'NPO法人ひまわり',
        'tenant_code': 'himawari-npo',
        'notification_email': 'saiyou@himawari-npo.or.jp',
    },
    'departments': ['就労支援事業部', '生活支援事業部', '相談支援事業部', '法人本部'],
    'staff': [
        {'id': 'u001', 'name': '川村 直人', 'role': '採用担当', 'dept': '法人本部'},
        {'id': 'u002', 'name': '吉川 恵', 'role': 'サービス管理責任者（就労A型）', 'dept': '就労支援事業部'},
        {'id': 'u003', 'name': '野村 健', 'role': 'サービス管理責任者（GH）', 'dept': '生活支援事業部'},
        {'id': 'u004', 'name': '竹内 真理', 'role': '相談支援専門員', 'dept': '相談支援事業部'},
        {'id': 'u005', 'name': '浜田 拓郎', 'role': '理事長', 'dept': '法人本部'},
    ],
    'jobs': [
        {'code': 'S-001', 'title': '生活支援員（就労継続支援A型）', 'dept': '就労支援事業部', 'type': 'full_time', 'salary_min': 250, 'salary_max': 320, 'location': '埼玉県さいたま市', 'headcount': 3},
        {'code': 'S-002', 'title': '生活支援員（就労継続支援B型）', 'dept': '就労支援事業部', 'type': 'full_time', 'salary_min': 240, 'salary_max': 300, 'location': '埼玉県さいたま市', 'headcount': 2},
        {'code': 'S-003', 'title': 'サービス管理責任者', 'dept': '就労支援事業部', 'type': 'full_time', 'salary_min': 350, 'salary_max': 450, 'location': '埼玉県さいたま市', 'headcount': 1},
        {'code': 'S-004', 'title': '世話人（グループホーム）', 'dept': '生活支援事業部', 'type': 'part_time', 'salary_min': None, 'salary_max': None, 'location': '埼玉県川口市', 'headcount': 4},
        {'code': 'S-005', 'title': '生活支援員（グループホーム）', 'dept': '生活支援事業部', 'type': 'full_time', 'salary_min': 260, 'salary_max': 330, 'location': '埼玉県川口市', 'headcount': 2},
        {'code': 'S-006', 'title': '相談支援専門員', 'dept': '相談支援事業部', 'type': 'full_time', 'salary_min': 320, 'salary_max': 400, 'location': '埼玉県さいたま市', 'headcount': 1},
        {'code': 'S-007', 'title': '就労支援員（就労移行支援）', 'dept': '就労支援事業部', 'type': 'full_time', 'salary_min': 280, 'salary_max': 350, 'location': '埼玉県さいたま市', 'headcount': 2},
        {'code': 'S-008', 'title': '職業指導員（就労継続支援A型）', 'dept': '就労支援事業部', 'type': 'full_time', 'salary_min': 270, 'salary_max': 340, 'location': '埼玉県さいたま市', 'headcount': 2},
        {'code': 'S-009', 'title': '児童発達支援管理責任者', 'dept': '生活支援事業部', 'type': 'full_time', 'salary_min': 350, 'salary_max': 450, 'location': '埼玉県川越市', 'headcount': 1},
        {'code': 'S-010', 'title': '児童指導員', 'dept': '生活支援事業部', 'type': 'full_time', 'salary_min': 260, 'salary_max': 330, 'location': '埼玉県川越市', 'headcount': 3},
    ],
    'candidates': [
        {'name': '星野 和彦', 'kana': 'ほしの かずひこ', 'gender': 'male', 'age': 35, 'exp': 8, 'current': '就労継続支援B型', 'position': '生活支援員', 'qual': '社会福祉士,精神保健福祉士', 'desired': 'サービス管理責任者'},
        {'name': '池田 沙織', 'kana': 'いけだ さおり', 'gender': 'female', 'age': 32, 'exp': 7, 'current': '相談支援事業所', 'position': '相談支援専門員', 'qual': '相談支援専門員,社会福祉士', 'desired': '相談支援'},
        {'name': '荒木 大輔', 'kana': 'あらき だいすけ', 'gender': 'male', 'age': 40, 'exp': 12, 'current': '就労移行支援', 'position': 'サービス管理責任者', 'qual': 'サービス管理責任者,精神保健福祉士', 'desired': 'サービス管理責任者'},
        {'name': '宮本 千晴', 'kana': 'みやもと ちはる', 'gender': 'female', 'age': 28, 'exp': 4, 'current': 'グループホーム', 'position': '生活支援員', 'qual': '介護福祉士', 'desired': '生活支援員'},
        {'name': '島田 誠', 'kana': 'しまだ まこと', 'gender': 'male', 'age': 45, 'exp': 18, 'current': '児童発達支援', 'position': '児童発達支援管理責任者', 'qual': '児童発達支援管理責任者,保育士', 'desired': '児発管'},
        {'name': '大塚 理恵', 'kana': 'おおつか りえ', 'gender': 'female', 'age': 30, 'exp': 5, 'current': '放課後等デイサービス', 'position': '児童指導員', 'qual': '児童指導員任用資格,保育士', 'desired': '児童指導員'},
        {'name': '平野 慎太郎', 'kana': 'ひらの しんたろう', 'gender': 'male', 'age': 33, 'exp': 6, 'current': '就労継続支援A型', 'position': '職業指導員', 'qual': '職業指導員', 'desired': '就労支援'},
        {'name': '菊池 美香', 'kana': 'きくち みか', 'gender': 'female', 'age': 27, 'exp': 3, 'current': '知的障害者施設', 'position': '生活支援員', 'qual': '社会福祉主事', 'desired': '生活支援員'},
        {'name': '和田 剛', 'kana': 'わだ つよし', 'gender': 'male', 'age': 38, 'exp': 10, 'current': '精神科デイケア', 'position': '精神保健福祉士', 'qual': '精神保健福祉士', 'desired': '相談支援'},
        {'name': '秋山 絵里', 'kana': 'あきやま えり', 'gender': 'female', 'age': 25, 'exp': 2, 'current': '障害者就労支援', 'position': '就労支援員', 'qual': '社会福祉士', 'desired': '就労支援'},
        {'name': '久保田 健一', 'kana': 'くぼた けんいち', 'gender': 'male', 'age': 42, 'exp': 15, 'current': '障害者支援施設', 'position': 'サービス管理責任者', 'qual': 'サービス管理責任者,介護福祉士', 'desired': 'サービス管理責任者'},
        {'name': '土井 奈々', 'kana': 'どい なな', 'gender': 'female', 'age': 29, 'exp': 4, 'current': 'グループホーム', 'position': '世話人', 'qual': '', 'desired': '世話人'},
        {'name': '橋本 翔', 'kana': 'はしもと しょう', 'gender': 'male', 'age': 26, 'exp': 2, 'current': '放課後等デイサービス', 'position': '指導員', 'qual': '児童指導員任用資格', 'desired': '児童指導員'},
        {'name': '川島 真由', 'kana': 'かわしま まゆ', 'gender': 'female', 'age': 34, 'exp': 8, 'current': '相談支援事業所', 'position': '主任相談員', 'qual': '相談支援専門員,主任相談支援専門員', 'desired': '相談支援'},
        {'name': '松井 一郎', 'kana': 'まつい いちろう', 'gender': 'male', 'age': 50, 'exp': 22, 'current': '社会福祉法人', 'position': '施設長', 'qual': '社会福祉士,管理者経験', 'desired': '管理職'},
        {'name': '中島 恵美', 'kana': 'なかじま えみ', 'gender': 'female', 'age': 31, 'exp': 6, 'current': '就労継続支援B型', 'position': '生活支援員', 'qual': '介護福祉士', 'desired': '生活支援員,A型'},
        {'name': '田村 修', 'kana': 'たむら おさむ', 'gender': 'male', 'age': 36, 'exp': 9, 'current': '就労移行支援', 'position': '就労支援員', 'qual': '就労支援員,キャリアコンサルタント', 'desired': '就労支援'},
        {'name': '早川 由紀子', 'kana': 'はやかわ ゆきこ', 'gender': 'female', 'age': 44, 'exp': 16, 'current': '児童発達支援', 'position': '保育士', 'qual': '保育士,幼稚園教諭', 'desired': '児童指導員'},
        {'name': '小野 達也', 'kana': 'おの たつや', 'gender': 'male', 'age': 24, 'exp': 1, 'current': '福祉系大学卒業', 'position': '', 'qual': '社会福祉士（新卒）', 'desired': '生活支援員,新卒'},
        {'name': '本田 真奈美', 'kana': 'ほんだ まなみ', 'gender': 'female', 'age': 37, 'exp': 10, 'current': 'グループホーム', 'position': '管理者', 'qual': 'サービス管理責任者,介護福祉士', 'desired': 'サービス管理責任者'},
    ],
    'sources': ['福祉のお仕事', 'LITALICO仕事ナビ', 'ジョブメドレー', 'ハローワーク', 'Indeed', '自社採用サイト'],
}


def generate_id(prefix, num, length=4):
    return f"{prefix}{str(num).zfill(length)}-0000-0000-0000-{'0' * 11}{str(num).zfill(1)}"


def generate_date(days_ago_min=1, days_ago_max=60):
    days_ago = random.randint(days_ago_min, days_ago_max)
    dt = datetime.now() - timedelta(days=days_ago)
    return dt.strftime('%Y-%m-%dT%H:%M:%S+09:00')


def generate_birth_date(age):
    year = datetime.now().year - age
    month = random.randint(1, 12)
    day = random.randint(1, 28)
    return f"{year}-{month:02d}-{day:02d}"


def generate_candidates_csv(config, output_path):
    """候補者CSVを生成"""
    headers = [
        'id', 'name', 'name_kana', 'email', 'phone', 'birth_date', 'gender',
        'postal_code', 'address', 'current_company', 'current_position',
        'years_of_experience', 'desired_salary_min', 'desired_salary_max',
        'desired_job_types', 'skills', 'qualifications', 'education',
        'source', 'status', 'notes', 'registered_by_id', 'created_at', 'updated_at'
    ]

    statuses = ['new', 'active', 'in_process', 'hired', 'rejected', 'withdrawn', 'inactive']
    educations = ['高卒', '専門卒', '短大卒', '大卒', '大学院卒']

    rows = []
    for i, c in enumerate(config['candidates'], 1):
        salary_min = 300 + c['exp'] * 10 + random.randint(-20, 20)
        salary_max = salary_min + random.randint(50, 150)
        status = random.choice(statuses[:4]) if i <= 15 else random.choice(statuses)
        source = random.choice(config['sources'])
        staff = random.choice(config['staff'])

        row = [
            generate_id('c', i),
            c['name'],
            c['kana'],
            f"candidate{i}@example.com",
            f"090-{random.randint(1000,9999)}-{random.randint(1000,9999)}",
            generate_birth_date(c['age']),
            c['gender'],
            f"{random.randint(100,999)}-{random.randint(1000,9999)}",
            f"東京都{random.choice(['世田谷区', '新宿区', '渋谷区', '目黒区', '港区'])}",
            c['current'],
            c['position'],
            c['exp'],
            salary_min,
            salary_max,
            c['desired'],
            c['qual'],
            c['qual'],
            random.choice(educations),
            source,
            status,
            '',
            staff['id'],
            generate_date(30, 90),
            generate_date(1, 30),
        ]
        rows.append(row)

    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    return len(rows)


def generate_jobs_csv(config, output_path):
    """求人CSVを生成"""
    headers = [
        'id', 'title', 'department', 'employment_type', 'job_category',
        'description', 'requirements', 'preferred_skills', 'salary_min', 'salary_max',
        'work_location', 'work_hours', 'benefits', 'number_of_positions',
        'status', 'published_at', 'deadline', 'notes', 'created_by_id',
        'created_at', 'updated_at'
    ]

    rows = []
    for i, j in enumerate(config['jobs'], 1):
        staff = random.choice(config['staff'])
        status = 'published' if i <= 7 else random.choice(['draft', 'published', 'paused'])

        row = [
            generate_id('j', i),
            j['title'],
            j['dept'],
            j['type'],
            j['title'].split('（')[0] if '（' in j['title'] else j['title'],
            f"{j['title']}の業務全般を担当していただきます。",
            "詳細は面接時にご説明します。",
            "経験者優遇",
            j['salary_min'] if j['salary_min'] else '',
            j['salary_max'] if j['salary_max'] else '',
            j['location'],
            '8:30-17:30（シフト制あり）' if 'part' not in j['type'] else 'シフト制',
            '社会保険完備、交通費支給、賞与年2回',
            j['headcount'],
            status,
            generate_date(30, 60) if status == 'published' else '',
            '2025-03-31',
            '',
            staff['id'],
            generate_date(60, 90),
            generate_date(1, 30),
        ]
        rows.append(row)

    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    return len(rows)


def generate_applications_csv(config, num_apps, output_path):
    """応募CSVを生成"""
    headers = [
        'id', 'candidate_id', 'candidate_name', 'job_id', 'job_title',
        'status', 'source', 'applied_at', 'evaluation_score', 'evaluation_notes',
        'offer_salary', 'offer_made_at', 'offer_deadline', 'offer_notes',
        'joined_at', 'notes', 'registered_by_id', 'created_at', 'updated_at'
    ]

    statuses = [
        'new', 'document_screening', 'document_passed', 'document_rejected',
        'interview_scheduled', 'interviewing', 'offer_pending', 'offer_made',
        'offer_accepted', 'offer_declined', 'rejected', 'withdrawn', 'on_hold'
    ]

    rows = []
    used_combinations = set()

    for i in range(1, num_apps + 1):
        # 重複しない候補者-求人の組み合わせを選ぶ
        while True:
            cand_idx = random.randint(1, len(config['candidates']))
            job_idx = random.randint(1, len(config['jobs']))
            combo = (cand_idx, job_idx)
            if combo not in used_combinations:
                used_combinations.add(combo)
                break

        cand = config['candidates'][cand_idx - 1]
        job = config['jobs'][job_idx - 1]
        status = random.choice(statuses[:8]) if i <= 15 else random.choice(statuses)
        staff = random.choice(config['staff'])
        source = random.choice(config['sources'])

        eval_score = random.randint(2, 5) if status not in ['new', 'document_screening'] else ''
        offer_salary = job['salary_min'] + random.randint(0, 50) if status in ['offer_made', 'offer_accepted'] and job['salary_min'] else ''

        row = [
            generate_id('a', i),
            generate_id('c', cand_idx),
            cand['name'],
            generate_id('j', job_idx),
            job['title'],
            status,
            source,
            generate_date(1, 60),
            eval_score,
            '面接評価コメント' if eval_score else '',
            offer_salary,
            generate_date(1, 14) if offer_salary else '',
            '2025-01-31' if offer_salary else '',
            '',
            '',
            '',
            staff['id'],
            generate_date(1, 60),
            generate_date(1, 7),
        ]
        rows.append(row)

    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    return len(rows)


def generate_interviews_csv(config, num_interviews, output_path):
    """面接CSVを生成"""
    headers = [
        'id', 'application_id', 'candidate_name', 'job_title',
        'interview_type', 'round_number', 'scheduled_at', 'duration_minutes',
        'location', 'meeting_url', 'interviewer_ids', 'interviewer_names',
        'status', 'result', 'feedback', 'score', 'notes', 'created_by_id',
        'created_at', 'updated_at'
    ]

    interview_types = ['in_person', 'video', 'phone', 'final', 'technical']
    statuses = ['scheduled', 'confirmed', 'completed', 'cancelled', 'no_show']
    results = ['passed', 'failed', 'pending']

    rows = []
    for i in range(1, num_interviews + 1):
        app_idx = random.randint(1, 15)  # 応募の一部のみ
        cand_idx = (app_idx - 1) % len(config['candidates']) + 1
        job_idx = (app_idx - 1) % len(config['jobs']) + 1

        cand = config['candidates'][cand_idx - 1]
        job = config['jobs'][job_idx - 1]
        staff = random.choice(config['staff'])

        itype = random.choice(interview_types)
        status = random.choice(statuses[:3])
        result = random.choice(results) if status == 'completed' else 'pending'

        row = [
            generate_id('i', i),
            generate_id('a', app_idx),
            cand['name'],
            job['title'],
            itype,
            random.randint(1, 3),
            generate_date(-14, 14).replace('T', 'T' + f"{random.randint(9,17):02d}:00:00"),
            random.choice([30, 45, 60, 90]),
            f"{config['tenant']['company_name']} 会議室" if itype == 'in_person' else '',
            'https://meet.google.com/xxx-yyyy-zzz' if itype == 'video' else '',
            staff['id'],
            staff['name'],
            status,
            result,
            '面接フィードバック' if result != 'pending' else '',
            random.randint(3, 5) if result == 'passed' else random.randint(1, 3) if result == 'failed' else '',
            '',
            staff['id'],
            generate_date(7, 30),
            generate_date(1, 7),
        ]
        rows.append(row)

    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    return len(rows)


def generate_settings_csv(config, output_path):
    """設定CSVを生成"""
    headers = ['key', 'value', 'description']

    rows = [
        ['company_name', config['tenant']['company_name'], '会社名'],
        ['tenant_code', config['tenant']['tenant_code'], 'テナントコード'],
        ['timezone', 'Asia/Tokyo', 'タイムゾーン'],
        ['date_format', 'YYYY-MM-DD', '日付形式'],
        ['currency', 'JPY', '通貨'],
        ['default_interview_duration', '60', 'デフォルト面接時間（分）'],
        ['notification_email', config['tenant']['notification_email'], '通知先メールアドレス'],
    ]

    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    return len(rows)


def create_excel_from_csvs(pattern_name, pattern_dir):
    """CSVからExcelを生成"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    sheets = [
        ('候補者', 'candidates.csv'),
        ('求人', 'jobs.csv'),
        ('応募', 'applications.csv'),
        ('面接', 'interviews.csv'),
        ('設定', 'settings.csv'),
    ]

    wb = Workbook()

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for idx, (sheet_name, csv_file) in enumerate(sheets):
        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        csv_path = pattern_dir / csv_file
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, start=1):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    if row_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment

        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                        max_length = max(max_length, length)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(50, max(8, max_length + 2))

        ws.freeze_panes = 'A2'

    excel_path = pattern_dir / f'{pattern_name}.xlsx'
    wb.save(excel_path)
    return excel_path


def generate_pattern(config, pattern_name, pattern_dir):
    """1パターン分のデータを生成"""
    pattern_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n=== {pattern_name} ===")
    print(f"Company: {config['tenant']['company_name']}")

    # CSVファイル生成
    n = generate_candidates_csv(config, pattern_dir / 'candidates.csv')
    print(f"  candidates.csv: {n} records")

    n = generate_jobs_csv(config, pattern_dir / 'jobs.csv')
    print(f"  jobs.csv: {n} records")

    n = generate_applications_csv(config, 20, pattern_dir / 'applications.csv')
    print(f"  applications.csv: {n} records")

    n = generate_interviews_csv(config, 12, pattern_dir / 'interviews.csv')
    print(f"  interviews.csv: {n} records")

    n = generate_settings_csv(config, pattern_dir / 'settings.csv')
    print(f"  settings.csv: {n} records")

    # Excel生成
    excel_path = create_excel_from_csvs(pattern_name, pattern_dir)
    print(f"  Excel: {excel_path.name}")

    return excel_path


def main():
    print("=" * 60)
    print("3 Industry Patterns Test Data Generator")
    print("=" * 60)

    patterns = [
        (KAIGO_CONFIG, 'Pattern1_Kaigo', 'kaigo'),
        (IRYO_CONFIG, 'Pattern2_Iryo', 'iryo'),
        (SHOGAI_CONFIG, 'Pattern3_Shogai', 'shogai'),
    ]

    excel_files = []
    for config, pattern_name, subdir in patterns:
        pattern_dir = OUTPUT_DIR / subdir
        excel_path = generate_pattern(config, pattern_name, pattern_dir)
        excel_files.append(excel_path)

    print("\n" + "=" * 60)
    print("Generation Complete!")
    print("=" * 60)
    print("\nExcel files created:")
    for path in excel_files:
        print(f"  - {path}")


if __name__ == '__main__':
    main()
