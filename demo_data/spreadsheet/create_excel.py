"""CSVファイルをExcelファイルに変換"""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 作業ディレクトリ
base_dir = Path(__file__).parent

# シート定義（順番通り）
sheets = [
    ('候補者', '候補者.csv'),
    ('求人', '求人.csv'),
    ('応募', '応募.csv'),
    ('面接', '面接.csv'),
    ('設定', '設定.csv'),
]

# Excelファイル作成
wb = Workbook()

# スタイル定義
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for idx, (sheet_name, csv_file) in enumerate(sheets):
    # シート作成（最初のシートはデフォルトを使用）
    if idx == 0:
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(title=sheet_name)

    # CSV読み込み
    csv_path = base_dir / csv_file
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # ヘッダー行のスタイル
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

    # 列幅を自動調整
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    # 日本語文字は幅2として計算
                    length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                    max_length = max(max_length, length)
        # 最大幅50、最小幅8
        ws.column_dimensions[column_letter].width = min(50, max(8, max_length + 2))

    # ヘッダー行を固定
    ws.freeze_panes = 'A2'

    print(f"[OK] {sheet_name}: {ws.max_row - 1} records")

# 保存
output_path = base_dir / 'ATS_TestData.xlsx'
wb.save(output_path)
print(f"\nExcel file created: {output_path}")
